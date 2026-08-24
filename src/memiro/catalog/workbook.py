"""Книга расчёта: тот же прайс, но в таблице, которую владелец крутит.

Владельцу нужно щупать цену до того, как она появится на витрине:
поднять ставку, поменять порог, посмотреть, что стало с зеркалом
2300 × 900. В админке это правка боевых данных, и смотреть на
последствия приходится уже после них.

Книга собирается из тех же данных, что и витрина, и повторяет ту же
арифметику формулами таблицы: единицы расхода, коэффициент формы,
пороги и округление (ADR-0007). Расхождение книги с сайтом — дефект,
и лист «Сверка» существует ровно затем, чтобы его было видно: итоги
там считает движок сайта, а рядом ту же конфигурацию считает книга.

Файлом в репозитории книга не лежит: ставки в ней протухают в тот же
день, когда владелец правит справочник. Лежит генератор, а книга
берётся командой `pricing_workbook` из живой базы.

Один лист данными не подкреплён — «Наценка за размер» (ADR-0010).
Решение принято, кода ещё нет, и признак у значения тоже: колонка
«умножается наценкой за размер» в справочнике пустая, владелец
проставляет её в книге сам. Когда признак появится в модели,
заполнять её станет `_dictionary_sheet()` — одной строкой.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from memiro import pricing
from . import tariffs
from .models import Attribute, AttributeValue, PricingSettings

if TYPE_CHECKING:
    from collections.abc import Sequence

    from openpyxl.worksheet.worksheet import Worksheet

    from .models import Category, Product

# Ступени наценки за размер: решение владельца из ADR-0010, которого
# ещё нет в базе. Книга показывает его как заготовку — владелец правит
# пороги прямо в ней. Когда таблица появится в админке, значения
# приедут оттуда, а константа уйдёт
PLANNED_SURCHARGE: tuple[tuple[int, Decimal], ...] = ((2200, Decimal("1.25")),)

# Размеры листа «Сверка»: маленькое (упирается в минимальную площадь),
# обычное и крупное. По ним видно все три нижних правила расчёта сразу
CHECK_SIZES: tuple[tuple[int, int], ...] = (
    (400, 300),
    (800, 600),
    (1200, 700),
    (2300, 900),
)

INK = "1A2422"
EDGE = "0E6656"
BRASS = "8A5A28"
FAINT = "8A9793"
SOFT_INK = "55635F"

HEAD = PatternFill("solid", fgColor="E2EFEB")
INPUT = PatternFill("solid", fgColor="FFF6E3")
SOFT = PatternFill("solid", fgColor="F4F7F6")
TOTAL = PatternFill("solid", fgColor="F6EDE2")
THIN = Side(style="thin", color="C9D5D1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.##" ₽"'
MONEY_ROUND = '#,##0" ₽"'
FACTOR_FORMAT = '0.0"×"'

# Единицы расхода словами: движок знает их значениями, а книга —
# подписями, и подписи эти стоят в формулах. Разъехавшись, они
# оставили бы книгу с нулями там, где у сайта цена
UNIT_LABELS = {
    pricing.Unit.PIECE: "за штуку",
    pricing.Unit.LINEAR_METER: "за пог. м",
    pricing.Unit.SQUARE_METER: "за м²",
    pricing.Unit.FACTOR: "коэффициент",
}
FREE_LABEL = "—"

DICTIONARY = "Справочник"
SURCHARGE = "Наценка за размер"
SETTINGS_SHEET = "Параметры расчёта"

# Первая строка данных на каждом листе: под заголовком, подзаголовком
# и шапкой таблицы
DATA_START = 5


@dataclass(frozen=True, slots=True)
class Line:
    """Строка расчёта: чем крутят и что стоит у товара по умолчанию."""

    attribute: Attribute
    default: AttributeValue | None
    quantity: Decimal


@dataclass(frozen=True, slots=True)
class Layout:
    """Где что лежит на листе «Расчёт».

    Строк у таблицы столько, сколько у категории атрибутов, и все
    номера ниже неё поэтому считаются, а не пишутся числами.
    """

    lines: int

    @property
    def first(self) -> int:
        return 15

    @property
    def last(self) -> int:
        return self.first + self.lines - 1

    @property
    def totals(self) -> int:
        return self.last + 3

    @property
    def order_minimum(self) -> int:
        return self.totals + 1

    @property
    def rounded(self) -> int:
        return self.totals + 2

    @property
    def within_limits(self) -> int:
        return self.totals + 3

    @property
    def shown(self) -> int:
        return self.totals + 4


def build(
    category: Category,
    *,
    product: Product | None = None,
) -> Workbook:
    """Книга расчёта по данным категории — и товара, если он назван."""
    attributes = _attributes_of(category)
    lines = _lines(attributes, product)
    ranges = _dictionary_ranges(attributes)
    settings = PricingSettings.objects.first()
    workbook = Workbook()
    workbook.remove(workbook.active)
    _guide_sheet(workbook.create_sheet("Как пользоваться"))
    _calculation_sheet(workbook.create_sheet("Расчёт"), lines, ranges)
    _dictionary_sheet(workbook.create_sheet(DICTIONARY), attributes, ranges)
    _attributes_sheet(workbook.create_sheet("Атрибуты"), attributes)
    _settings_sheet(workbook.create_sheet(SETTINGS_SHEET), settings)
    _surcharge_sheet(workbook.create_sheet(SURCHARGE))
    _check_sheet(workbook.create_sheet("Сверка"), product, settings)
    return workbook


def _attributes_of(category: Category) -> list[Attribute]:
    """Атрибуты категории со справочником — одним запросом."""
    return list(
        Attribute.objects.filter(category=category)
        .prefetch_related("values", "parents")
        .order_by("order", "name")
    )


def _lines(
    attributes: Sequence[Attribute], product: Product | None
) -> list[Line]:
    """Строки расчёта: платные виды атрибутов и умолчания товара.

    «Да/нет» сюда не попадает: значения справочника у него нет, а
    значит нет ни единицы расхода, ни тарифа — в цену такой атрибут
    не входит вовсе (`tariffs.declared_values()`).
    """
    chosen, counted = _defaults_of(product)
    return [
        Line(
            attribute=attribute,
            default=chosen.get(attribute.pk) or _first_value(attribute),
            quantity=counted.get(attribute.pk, Decimal(1)),
        )
        for attribute in attributes
        if attribute.kind != Attribute.Kind.BOOLEAN
    ]


def _defaults_of(
    product: Product | None,
) -> tuple[dict[int, AttributeValue], dict[int, Decimal]]:
    """Разметка товара: что выбрано и сколько раз."""
    if product is None:
        return {}, {}
    rows = product.attribute_values.all()
    return (
        {
            row.attribute_id: row.value_option
            for row in rows
            if row.value_option is not None
        },
        {
            row.attribute_id: row.value_number
            for row in rows
            if row.value_number is not None
        },
    )


def _first_value(attribute: Attribute) -> AttributeValue | None:
    """Чем открывается список, когда товар не назван."""
    return next(iter(attribute.values.all()), None)


def _unit_label(value: AttributeValue) -> str:
    """Единица расхода подписью — или прочерк у бесплатного."""
    if not value.unit:
        return FREE_LABEL
    return UNIT_LABELS[pricing.Unit(value.unit)]


def _key(attribute_name: str, value_name: str) -> str:
    """Ключ строки справочника: атрибут и значение вместе.

    Одним значением не обойтись: «Серебро» бывает и полотном, и цветом
    рамы, и по одному слову расчёт нашёл бы не ту ставку.
    """
    return f"{attribute_name} | {value_name}"


def _title(sheet: Worksheet, cell: str, text: str, size: int = 16) -> None:
    sheet[cell] = text
    sheet[cell].font = Font(bold=True, size=size, color=INK)


def _subtitle(sheet: Worksheet, cell: str, text: str, span: str) -> None:
    sheet[cell] = text
    sheet[cell].font = Font(size=11, color=SOFT_INK, italic=True)
    sheet[cell].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(span)
    sheet.row_dimensions[sheet[cell].row].height = 30


def _header(sheet: Worksheet, row: int, titles: Sequence[str]) -> None:
    for index, title in enumerate(titles, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.font = Font(bold=True, size=10, color=EDGE)
        cell.fill = HEAD
        cell.border = BOX
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[row].height = 30


def _widths(sheet: Worksheet, sizes: dict[str, int]) -> None:
    for column, width in sizes.items():
        sheet.column_dimensions[column].width = width


def _tip(sheet: Worksheet, cell: str, text: str) -> None:
    """Заметка при наведении — в Google Таблицах это «примечание»."""
    comment = Comment(text, "Memiro")
    comment.width = 320
    comment.height = 20 + 13 * (1 + len(text) // 45)
    sheet[cell].comment = comment


def _boxed(sheet: Worksheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        sheet.cell(row=row, column=column).border = BOX


def _dictionary_ranges(
    attributes: Sequence[Attribute],
) -> dict[int, tuple[int, int]]:
    """Где на листе справочника лежат значения каждого атрибута.

    Списки выбора на «Расчёте» ссылаются на эти строки, а пишутся оба
    листа порознь — значит границы должны считаться одинаково и там,
    и там, из одного порядка атрибутов.
    """
    ranges: dict[int, tuple[int, int]] = {}
    row = DATA_START
    for attribute in attributes:
        values = list(attribute.values.all())
        if not values:
            continue
        ranges[attribute.pk] = (row, row + len(values) - 1)
        row += len(values)
    return ranges


def _lookup(column: str, key: str) -> str:
    """Поиск поля справочника по ключу «Атрибут | Значение»."""
    return (
        f"IFERROR(INDEX({DICTIONARY}!${column}:${column},"
        f'MATCH({key},{DICTIONARY}!$H:$H,0)),"")'
    )


def _guide_sheet(sheet: Worksheet) -> None:
    """Первый лист: что это и куда жать."""
    _title(sheet, "A1", "Как пользоваться книгой", 18)
    sheet.sheet_view.showGridLines = False
    _widths(sheet, {"A": 4, "B": 30, "C": 100})
    blocks = (
        (
            "Что это",
            (
                "Точная копия расчёта, который сайт делает на карточке "
                "товара. Собрана из живых данных админки командой "
                "pricing_workbook."
            ),
        ),
        (
            "Жёлтые ячейки",
            (
                "Всё, что вводится руками. Остальное — формулы, их лучше "
                "не трогать. Наведите на заголовок: там подсказка."
            ),
        ),
        ("", ""),
        (
            "Лист «Расчёт»",
            (
                "Собираете зеркало: размеры сверху, характеристики "
                "выпадающими списками. «Умолчание товара» — то, чем "
                "размечен товар; «Выбор покупателя» — то, что он крутит. "
                "Из разницы получается доплата, которую он видит."
            ),
        ),
        (
            "Лист «Справочник»",
            (
                "Все цены. Единица расхода решает, на что умножится "
                "тариф: на площадь, на периметр или на штуку."
            ),
        ),
        (
            "Лист «Наценка за размер»",
            (
                "Ступени для крупных изделий. В админке этой таблицы "
                "пока нет — решение записано в ADR-0010, и книга нужна, "
                "чтобы покрутить пороги до того, как их напишут в коде."
            ),
        ),
        (
            "Лист «Сверка»",
            (
                "Итоги, посчитанные движком сайта. Соберите то же "
                "зеркало на «Расчёте» — числа обязаны совпасть."
            ),
        ),
        ("", ""),
        (
            "Три единицы расхода",
            (
                "За квадратный метр — полотно: расходуется листом. За "
                "погонный метр — рама, кромка, лента: идут по контуру. "
                "За штуку — кнопка, подогрев, крепление, вырез."
            ),
        ),
        (
            "Коэффициент формы",
            (
                "Умножает только то, что режется по контуру. На ленту не "
                "действует: на круге она не дороже, а короче."
            ),
        ),
        (
            "Порядок сложения",
            (
                "Статьи складываются, итог поднимается до минимальной "
                "суммы заказа и округляется вверх до шага округления."
            ),
        ),
        (
            "Доплата покупателю",
            (
                "Считается по точным статьям, до порога и до округления. "
                "Разница двух итогов на пороге соврала бы."
            ),
        ),
    )
    row = 3
    for head, text in blocks:
        if not head:
            row += 1
            continue
        title = sheet.cell(row=row, column=2, value=head)
        title.font = Font(bold=True, size=11, color=EDGE)
        title.alignment = Alignment(vertical="top")
        body = sheet.cell(row=row, column=3, value=text)
        body.font = Font(size=11, color=INK)
        body.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 15 * (1 + len(text) // 95)
        row += 1


def _calculation_sheet(
    sheet: Worksheet,
    lines: Sequence[Line],
    ranges: dict[int, tuple[int, int]],
) -> None:
    """Рабочий лист: размеры сверху, статьи в таблице, итог внизу."""
    layout = Layout(len(lines))
    _title(sheet, "A1", "Калькулятор цены зеркала", 18)
    _subtitle(
        sheet,
        "A2",
        "Жёлтые ячейки — ввод. Остальное считается по тарифам с листа "
        "«Справочник». Это тот же расчёт, что делает сайт.",
        "A2:L2",
    )
    _legend(sheet)
    _sizes_block(sheet, layout)
    _table(sheet, lines, layout, ranges)
    _totals_block(sheet, layout)
    _calculation_tips(sheet, layout)
    _widths(
        sheet,
        {
            "A": 24,
            "B": 19,
            "C": 19,
            "D": 8,
            "E": 15,
            "F": 12,
            "G": 10,
            "H": 9,
            "I": 10,
            "J": 14,
            "K": 19,
            "L": 46,
        },
    )
    for column in "MNOPQR":
        sheet.column_dimensions[column].hidden = True
    sheet.freeze_panes = f"A{layout.first}"
    sheet.sheet_view.showGridLines = False
    note = sheet.cell(
        row=layout.shown + 3,
        column=1,
        value="Скрытые столбцы с M по R: в них считается то же зеркало с "
        "умолчаниями товара — из разницы получается доплата.",
    )
    note.font = Font(size=10, italic=True, color=FAINT)


def _legend(sheet: Worksheet) -> None:
    """Две подписи с образцами цвета — что заполняется руками."""
    filled = sheet.cell(row=3, column=1, value="ЗАПОЛНЯЕТЕ ВЫ")
    filled.font = Font(bold=True, size=9, color=BRASS)
    filled.fill = INPUT
    filled.border = BOX
    filled.alignment = Alignment(horizontal="center")
    sheet.cell(
        row=3,
        column=2,
        value="Жёлтое — ввод: размеры, выбор из списков, количество.",
    ).font = Font(size=10, color=SOFT_INK)
    auto = sheet.cell(row=3, column=4, value="СЧИТАЕТСЯ САМО")
    auto.font = Font(bold=True, size=9, color=EDGE)
    auto.fill = SOFT
    auto.border = BOX
    auto.alignment = Alignment(horizontal="center")
    sheet.cell(
        row=3,
        column=5,
        value="Остальное — формулы. Наведите на заголовок: там подсказка.",
    ).font = Font(size=10, color=SOFT_INK)


def _sizes_block(sheet: Worksheet, layout: Layout) -> None:
    """Шаг 1: габариты и всё, что из них выводится."""
    _title(sheet, "A4", "Шаг 1.  Размеры изделия", 12)
    hint = sheet.cell(row=4, column=3, value="◀ вводите в жёлтых ячейках")
    hint.font = Font(size=10, italic=True, color=BRASS)
    rows = (
        ("Ширина, мм", 900, "#,##0", True),
        ("Высота, мм", 900, "#,##0", True),
        ("Площадь, м²", "=B5*B6/1000000", "0.000", False),
        (
            "Площадь в расчёте, м²",
            f"=MAX(B7,'{SETTINGS_SHEET}'!$B$5)",
            "0.000",
            False,
        ),
        ("Периметр, пог. м", "=2*(B5+B6)/1000", "0.000", False),
        (
            "Коэффициент формы",
            f"=PRODUCT(M{layout.first}:M{layout.last})",
            FACTOR_FORMAT,
            False,
        ),
        ("Наценка за размер", _surcharge_formula(), FACTOR_FORMAT, False),
    )
    for offset, (name, value, fmt, editable) in enumerate(rows):
        row = 5 + offset
        sheet.cell(row=row, column=1, value=name).border = BOX
        cell = sheet.cell(row=row, column=2, value=value)
        cell.border = BOX
        cell.number_format = fmt
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = INPUT if editable else SOFT


def _surcharge_formula() -> str:
    """Коэффициент из таблицы ступеней — по наибольшей стороне.

    Ступень ищется поиском по возрастанию: берётся последняя, чей
    порог изделие переросло. Размер ниже первого порога в таблицу не
    попадает вовсе, и `IFERROR` возвращает единицу — «наценки нет».
    Так же ведёт себя и пустая таблица (ADR-0010).
    """
    last = DATA_START + 19
    return (
        f"=IFERROR(LOOKUP(MAX(B5,B6),'{SURCHARGE}'!"
        f"$A${DATA_START}:$A${last},'{SURCHARGE}'!"
        f"$B${DATA_START}:$B${last}),1)"
    )


def _table(
    sheet: Worksheet,
    lines: Sequence[Line],
    layout: Layout,
    ranges: dict[int, tuple[int, int]],
) -> None:
    """Шаг 2: по строке на характеристику, с разложением по шагам."""
    _title(sheet, "A13", "Шаг 2.  Из чего складывается цена", 12)
    _column_pointers(sheet)
    _header(
        sheet,
        14,
        (
            "Атрибут",
            "Умолчание товара",
            "Выбор покупателя",
            "Кол-во",
            "Единица расхода",
            "Тариф, ₽",
            "Расход",
            "× форма",
            "× размер",
            "Сумма, ₽",
            "Доплата покупателю, ₽",
            "Подсказка",
        ),
    )
    for offset, line in enumerate(lines):
        _line_row(sheet, layout.first + offset, line, ranges)


def _column_pointers(sheet: Worksheet) -> None:
    """Подписи над столбцами: где смотреть, а где выбирать."""
    pointers = (
        (2, "как размечен товар", Font(size=9, italic=True, color=FAINT)),
        (
            3,
            "▼ ЗДЕСЬ ВЫБИРАЕТ ПОКУПАТЕЛЬ",
            Font(bold=True, size=9, color=BRASS),
        ),
        (10, "считается само", Font(size=9, italic=True, color=FAINT)),
    )
    for column, text, font in pointers:
        cell = sheet.cell(row=13, column=column, value=text)
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _line_row(
    sheet: Worksheet,
    row: int,
    line: Line,
    ranges: dict[int, tuple[int, int]],
) -> None:
    """Одна характеристика: выбор, тариф, расход и сумма."""
    name = line.attribute.name
    default = line.default.value if line.default else ""
    chosen = f'$A{row}&" | "&$C{row}'
    fallback = f'$A{row}&" | "&$B{row}'
    sheet.cell(row=row, column=1, value=name)
    sheet.cell(row=row, column=2, value=default).fill = SOFT
    choice = sheet.cell(row=row, column=3, value=default)
    choice.fill = INPUT
    choice.font = Font(bold=True)
    quantity = sheet.cell(row=row, column=4, value=line.quantity)
    quantity.fill = INPUT
    quantity.alignment = Alignment(horizontal="center")
    _priced_columns(sheet, row, chosen, visible=True)
    _priced_columns(sheet, row, fallback, visible=False)
    surcharge = sheet.cell(
        row=row,
        column=11,
        value=(
            f"=IF(INDEX(Атрибуты!$B:$B,MATCH($A{row},Атрибуты!$A:$A,0))"
            f'<>"да","—",IF(ROUND($J{row}-$R{row},2)=0,"—",'
            f"ROUND(ROUND($J{row}-$R{row},2),0)))"
        ),
    )
    surcharge.number_format = '+#,##0" ₽";-#,##0" ₽"'
    surcharge.font = Font(bold=True, color=BRASS)
    hint = sheet.cell(row=row, column=12, value=_line_hint(line))
    hint.font = Font(size=10, color=FAINT)
    hint.alignment = Alignment(wrap_text=True, vertical="center")
    _boxed(sheet, row, 12)
    sheet.row_dimensions[row].height = 26
    _dropdown(sheet, row, line, ranges)


def _priced_columns(
    sheet: Worksheet, row: int, key: str, *, visible: bool
) -> None:
    """Пять столбцов пути от значения до суммы.

    Считаются дважды: по выбору покупателя — на виду, и по умолчанию
    товара — в скрытых столбцах. Разница двух сумм и есть доплата, а
    считать её на пороге вычитанием готовых итогов нельзя (ADR-0007).
    """
    unit, rate, factor, size, total = (
        ("E", "F", "H", "I", "J") if visible else ("N", "O", "P", "Q", "R")
    )
    sheet[f"{unit}{row}"] = f"={_lookup('C', key)}"
    sheet[f"{rate}{row}"] = (
        f'=IF(OR(${unit}{row}="",${unit}{row}="{FREE_LABEL}"),"",'
        f"{_lookup('D', key)})"
    )
    sheet[f"{factor}{row}"] = f'=IF({_lookup("E", key)}="да",$B$10,1)'
    sheet[f"{size}{row}"] = f'=IF({_lookup("F", key)}="да",$B$11,1)'
    sheet[f"{total}{row}"] = (
        f'=IF(OR(${unit}{row}="",${unit}{row}="{FREE_LABEL}",'
        f'${unit}{row}="коэффициент",${rate}{row}=""),0,'
        f"${rate}{row}*$G{row}*$D{row}*${factor}{row}*${size}{row})"
    )
    if not visible:
        return
    consumption = sheet.cell(
        row=row,
        column=7,
        value=(
            f'=IF($E{row}="за м²",$B$8,IF($E{row}="за пог. м",$B$9,'
            f'IF($E{row}="за штуку",1,"")))'
        ),
    )
    consumption.number_format = "0.###"
    sheet[f"M{row}"] = f'=IF($E{row}="коэффициент",IF($F{row}="",1,$F{row}),1)'
    sheet[f"F{row}"].number_format = "#,##0.##"
    sheet[f"H{row}"].number_format = FACTOR_FORMAT
    sheet[f"I{row}"].number_format = FACTOR_FORMAT
    sheet[f"J{row}"].number_format = MONEY
    sheet[f"J{row}"].font = Font(bold=True)


def _line_hint(line: Line) -> str:
    """Короткое пояснение к строке — из вида атрибута и его значений."""
    if line.attribute.kind == Attribute.Kind.NUMBER:
        return "Число, а не выбор: ставьте количество в столбце «Кол-во»."
    units = {_unit_label(value) for value in line.attribute.values.all()} - {
        FREE_LABEL
    }
    if not units:
        return "Бесплатно: тарифа в справочнике нет."
    if units == {"коэффициент"}:
        return "Коэффициент: своей строки не даёт, умножает другие."
    return "Считается по единице расхода из справочника."


def _dropdown(
    sheet: Worksheet,
    row: int,
    line: Line,
    ranges: dict[int, tuple[int, int]],
) -> None:
    """Список значений этого атрибута — и только его.

    Чужое значение подставить нельзя: расчёт нашёл бы его в
    справочнике по ключу другого атрибута и молча дал бы не ту ставку.
    """
    bounds = ranges.get(line.attribute.pk)
    if bounds is None:
        return
    first, last = bounds
    validation = DataValidation(
        type="list",
        formula1=f"={DICTIONARY}!$B${first}:$B${last}",
        allow_blank=True,
    )
    validation.errorTitle = "Значение вне справочника"
    validation.error = "Такого значения у этого атрибута нет."
    sheet.add_data_validation(validation)
    validation.add(sheet.cell(row=row, column=2))
    validation.add(sheet.cell(row=row, column=3))


def _totals_block(sheet: Worksheet, layout: Layout) -> None:
    """Шаг 3: сумма статей, пороги, округление и строка витрины."""
    _title(sheet, f"A{layout.totals - 1}", "Шаг 3.  Итог", 12)
    rows = (
        (
            "Сумма статей",
            f"=SUM(J{layout.first}:J{layout.last})",
            MONEY,
            "Точные статьи, до порогов и округления.",
        ),
        (
            "Минимальная сумма заказа",
            f"='{SETTINGS_SHEET}'!$B$6",
            MONEY_ROUND,
            "Ниже неё итог не опускается.",
        ),
        (
            "Итог, округление вверх",
            (
                f"=CEILING(MAX(B{layout.totals},B{layout.order_minimum}),"
                f"'{SETTINGS_SHEET}'!$B$9)"
            ),
            MONEY_ROUND,
            "Максимум из двух строк выше, округлённый вверх.",
        ),
        ("Размер", _limits_formula(), None, "Сверяется с поворотом."),
        (
            "Что покажет сайт",
            _shown_formula(layout),
            None,
            "Ровно эта строка стоит на карточке товара.",
        ),
    )
    for offset, (name, value, fmt, hint) in enumerate(rows):
        row = layout.totals + offset
        title = sheet.cell(row=row, column=1, value=name)
        title.border = BOX
        cell = sheet.cell(row=row, column=2, value=value)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
        if fmt:
            cell.number_format = fmt
        if row in (layout.rounded, layout.shown):
            cell.fill = TOTAL
            size = 14 if row == layout.rounded else 12
            cell.font = Font(bold=True, size=size, color=BRASS)
            title.font = Font(bold=True, size=12)
            sheet.row_dimensions[row].height = 26
        else:
            cell.fill = SOFT
            cell.font = Font(bold=True)
        sheet.cell(row=row, column=3, value=hint).font = Font(
            size=10, color=FAINT
        )


def _limits_formula() -> str:
    """Помещается ли изделие в пределы производства."""
    return (
        f"=IF(AND(OR('{SETTINGS_SHEET}'!$B$7=0,"
        f"MAX(B5,B6)<='{SETTINGS_SHEET}'!$B$7),"
        f"OR('{SETTINGS_SHEET}'!$B$8=0,"
        f"MIN(B5,B6)<='{SETTINGS_SHEET}'!$B$8)),"
        '"в пределах производства","ЗА ПРЕДЕЛОМ производства")'
    )


def _shown_formula(layout: Layout) -> str:
    """Что стоит на карточке: цена или приглашение оставить заявку."""
    return (
        f'=IF(B{layout.within_limits}="ЗА ПРЕДЕЛОМ производства",'
        '"Цены нет — приглашение оставить заявку",'
        f"IF(B{layout.totals}=0,"
        '"Цены нет — ни одна характеристика не тарифицирована",'
        f'TEXT(B{layout.rounded},"#,##0")&" ₽"))'
    )


def _calculation_tips(sheet: Worksheet, layout: Layout) -> None:
    """Заметки при наведении — на размеры, шапку и итоги."""
    sizes = {
        "B5": "Ширина изделия в миллиметрах. На сайте её вводит покупатель.",
        "B6": "Высота изделия в миллиметрах.",
        "B7": "Ширина × высота ÷ 1 000 000.",
        "B8": "Та площадь, по которой считается полотно: не ниже "
        "минимальной из «Параметров расчёта».",
        "B9": "2 × (ширина + высота) ÷ 1000. По периметру считаются "
        "рама, кромка и лента.",
        "B10": "Из строки «Форма». Умножает только то, что режется по "
        "контуру.",
        "B11": "Из таблицы ступеней на листе «Наценка за размер». "
        "Умножает то, у чего в справочнике стоит её признак.",
    }
    headers = {
        "A14": "Характеристика изделия. Настройки — на листе «Атрибуты».",
        "B14": "Чем размечен товар в каталоге. От него считается доплата.",
        "C14": "Что выбрал покупатель. Нажмите на ячейку — появится "
        "стрелка списка.",
        "D14": "Сколько раз статья входит в изделие. Нужно числовым "
        "атрибутам: два выреза стоят вдвое.",
        "E14": "Чем статья меряется. Из «Справочника», здесь не правится.",
        "F14": "Цена за единицу. Меняется на листе «Справочник».",
        "G14": "Сколько единиц изделие расходует.",
        "H14": "Коэффициент формы, если он к этой статье применим.",
        "I14": "Наценка за размер, если она к этой статье применима.",
        "J14": "Тариф × расход × количество × коэффициенты.",
        "K14": "Разница с умолчанием товара — то, что видит "
        "покупатель. У неменяемых характеристик стоит «—».",
        "L14": "Короткое пояснение к строке.",
    }
    totals = {
        f"B{layout.totals}": "Сумма статей без порогов и округления.",
        f"B{layout.rounded}": "Итоговая цена изделия.",
        f"B{layout.shown}": "Сайт молчит о цене, если размер за "
        "пределом производства или ни одна характеристика не "
        "тарифицирована.",
    }
    for cell, text in (sizes | headers | totals).items():
        _tip(sheet, cell, text)


def _dictionary_sheet(
    sheet: Worksheet,
    attributes: Sequence[Attribute],
    ranges: dict[int, tuple[int, int]],
) -> None:
    """Справочник: все значения с тарифами — единственный источник цен."""
    _title(sheet, "A1", "Справочник: значения характеристик и тарифы")
    _subtitle(
        sheet,
        "A2",
        "Здесь живут все цены. Пустая единица и пустой тариф означают "
        "«бесплатно»: значение описывает изделие, но денег не стоит.",
        "A2:H2",
    )
    _header(
        sheet,
        4,
        (
            "Атрибут",
            "Значение",
            "Единица расхода",
            "Тариф, ₽",
            "Умножается\nформой",
            "Умножается\nнаценкой за размер",
            "Означает\nотсутствие",
            "Ключ (служебное)",
        ),
    )
    for attribute in attributes:
        bounds = ranges.get(attribute.pk)
        if bounds is None:
            continue
        for row, value in enumerate(attribute.values.all(), bounds[0]):
            _dictionary_row(sheet, row, attribute, value)
    _widths(
        sheet,
        {
            "A": 24,
            "B": 22,
            "C": 15,
            "D": 12,
            "E": 12,
            "F": 16,
            "G": 12,
            "H": 30,
        },
    )
    sheet.freeze_panes = f"A{DATA_START}"
    _dictionary_tips(sheet)


def _dictionary_row(
    sheet: Worksheet,
    row: int,
    attribute: Attribute,
    value: AttributeValue,
) -> None:
    """Строка справочника: значение, единица, тариф и признаки."""
    unit = _unit_label(value)
    sheet.cell(row=row, column=1, value=attribute.name)
    sheet.cell(row=row, column=2, value=value.value)
    label = sheet.cell(row=row, column=3, value=unit)
    if unit == FREE_LABEL:
        label.font = Font(color=FAINT)
    rate = sheet.cell(row=row, column=4, value=value.rate)
    rate.alignment = Alignment(horizontal="right")
    rate.number_format = FACTOR_FORMAT if unit == "коэффициент" else "#,##0.##"
    if value.rate is None:
        rate.fill = SOFT
    shape = sheet.cell(
        row=row, column=5, value="да" if value.scaled_by_shape else ""
    )
    shape.alignment = Alignment(horizontal="center")
    # Признака наценки за размер в модели ещё нет (ADR-0010): колонка
    # пустая, и владелец проставляет её в книге сам. Появится поле —
    # сюда встанет `value.scaled_by_size`, и книга начнёт повторять
    # админку и в этом
    size = sheet.cell(row=row, column=6, value="")
    size.alignment = Alignment(horizontal="center")
    size.fill = INPUT
    absence = sheet.cell(
        row=row, column=7, value="да" if value.marks_absence else ""
    )
    absence.alignment = Alignment(horizontal="center")
    key = sheet.cell(
        row=row, column=8, value=_key(attribute.name, value.value)
    )
    key.font = Font(color="B4BEBB", size=9)
    _boxed(sheet, row, 8)


def _dictionary_tips(sheet: Worksheet) -> None:
    """Подсказки к шапке справочника."""
    tips = {
        "C4": "За м² — полотно; за пог. м — рама, кромка, лента; за "
        "штуку — кнопка, подогрев, вырез; коэффициент — форма. "
        "Пусто означает «бесплатно».",
        "D4": "Цена за одну единицу. Главное поле: поменяли здесь — "
        "пересчиталась вся книга.",
        "E4": "«да» тому, что режется по контуру: полотну, раме, "
        "обработке кромки. Ленте не ставится: на круге она не "
        "дороже, а короче.",
        "F4": "Признака ещё нет в админке (ADR-0010). Проставьте «да» "
        "тому, что дорожает на крупном изделии: по решению "
        "владельца это одно полотно.",
        "G4": "«Без рамы», «без подсветки»: зависящие характеристики "
        "такое значение родителем не считают.",
        "H4": "Служебный столбец: по нему расчёт находит строку. Не трогайте.",
    }
    for cell, text in tips.items():
        _tip(sheet, cell, text)


def _attributes_sheet(
    sheet: Worksheet, attributes: Sequence[Attribute]
) -> None:
    """Настройки характеристики целиком — без тарифов."""
    _title(sheet, "A1", "Настройки характеристик")
    _subtitle(
        sheet,
        "A2",
        "Тарифы — на листе «Справочник». Здесь то, что относится к "
        "характеристике целиком: кто её выбирает и от чего она зависит.",
        "A2:E2",
    )
    _header(
        sheet,
        4,
        (
            "Атрибут",
            "Меняет\nпокупатель",
            "Тип значения",
            "Существует\nтолько при",
            "Строит\nфильтр",
        ),
    )
    editable_column = 2
    centred = (editable_column, 5)
    for row, attribute in enumerate(attributes, DATA_START):
        parents = ", ".join(parent.name for parent in attribute.parents.all())
        cells = (
            attribute.name,
            "да" if attribute.is_customer_editable else "нет",
            attribute.get_kind_display(),
            parents,
            "да" if attribute.is_filterable else "нет",
        )
        for column, text in enumerate(cells, start=1):
            cell = sheet.cell(row=row, column=column, value=text)
            cell.border = BOX
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if column in centred else "left",
            )
            if column == editable_column and text == "да":
                cell.font = Font(bold=True, color=EDGE)
    _widths(sheet, {"A": 26, "B": 14, "C": 18, "D": 26, "E": 10})
    sheet.freeze_panes = f"A{DATA_START}"
    _tip(
        sheet,
        "B4",
        "«да» — покупатель крутит это в калькуляторе и видит доплату. "
        "«нет» — характеристика описывает модель, но в расчёт входит.",
    )
    _tip(
        sheet,
        "D4",
        "Характеристики не бывает без родителя: температуры свечения "
        "не бывает у зеркала без подсветки.",
    )


def _settings_sheet(
    sheet: Worksheet, settings: PricingSettings | None
) -> None:
    """Границы расчёта — те же четыре числа, что в админке."""
    _title(sheet, "A1", "Параметры расчёта")
    _subtitle(
        sheet,
        "A2",
        "Четыре границы, общие для всего сайта, и шаг округления. "
        "Жёлтые ячейки меняете вы.",
        "A2:C2",
    )
    _header(sheet, 4, ("Параметр", "Значение", "Что делает"))
    rows = (
        (
            "Минимальная площадь, м²",
            settings.min_area_m2 if settings else Decimal(0),
            "0.000",
            (
                "Изделие меньше считается по ней: раскрой и обработка "
                "дешевле не становятся."
            ),
        ),
        (
            "Минимальная сумма заказа, ₽",
            settings.min_order_total if settings else 0,
            MONEY_ROUND,
            "Ниже этой суммы итог не опускается.",
        ),
        (
            "Наибольшая сторона, мм",
            settings.max_long_side_mm if settings else 0,
            '#,##0" мм"',
            (
                "Больше — производство не берёт, цена не показывается. "
                "0 означает «предела нет»."
            ),
        ),
        (
            "Вторая сторона, мм",
            settings.max_short_side_mm if settings else 0,
            '#,##0" мм"',
            "Сверяется с поворотом: порядок ввода не важен.",
        ),
        (
            "Шаг округления итога, ₽",
            pricing.ROUNDING_STEP,
            MONEY_ROUND,
            (
                "Итог округляется вверх до этого шага. В админке поля нет: "
                "шаг живёт в коде расчёта."
            ),
        ),
    )
    for row, (name, value, fmt, hint) in enumerate(rows, DATA_START):
        sheet.cell(row=row, column=1, value=name).border = BOX
        cell = sheet.cell(row=row, column=2, value=value)
        cell.border = BOX
        cell.number_format = fmt
        cell.fill = INPUT
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        note = sheet.cell(row=row, column=3, value=hint)
        note.border = BOX
        note.font = Font(size=10, color=SOFT_INK)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 30
    _widths(sheet, {"A": 30, "B": 16, "C": 74})
    if settings is None:
        warning = sheet.cell(
            row=DATA_START + 6,
            column=1,
            value="Параметры расчёта в админке не заведены — "
            "подставлены нули, то есть «границ нет».",
        )
        warning.font = Font(size=10, italic=True, color=BRASS)


def _surcharge_sheet(sheet: Worksheet) -> None:
    """Ступени наценки за размер — решение ADR-0010 до его кода."""
    _title(sheet, "A1", "Наценка за размер")
    _subtitle(
        sheet,
        "A2",
        "Ступени для крупных изделий: наибольшая сторона от — и "
        "коэффициент. В админке этой таблицы ещё нет (ADR-0010), "
        "и книга нужна, чтобы покрутить пороги до того, как их напишут.",
        "A2:C2",
    )
    _header(
        sheet,
        4,
        ("Наибольшая сторона от, мм", "Коэффициент", "Комментарий"),
    )
    for row, (side, factor) in enumerate(PLANNED_SURCHARGE, DATA_START):
        edge = sheet.cell(row=row, column=1, value=side)
        edge.number_format = '#,##0" мм"'
        edge.alignment = Alignment(horizontal="center")
        edge.fill = INPUT
        value = sheet.cell(row=row, column=2, value=factor)
        value.number_format = FACTOR_FORMAT
        value.alignment = Alignment(horizontal="center")
        value.fill = INPUT
        value.font = Font(bold=True)
        _boxed(sheet, row, 3)
    rules = (
        (
            "Ступени идут по возрастанию порога; изделие получает "
            "последнюю, чей порог оно переросло."
        ),
        (
            "Наибольшая сторона, а не площадь: режут из листа, и полоса "
            "3000 × 200 в лист не ложится."
        ),
        (
            "Пустая таблица означает, что наценки нет. Строка с "
            "коэффициентом 1,0 для этого не нужна."
        ),
        (
            "Умножает только то, чему на «Справочнике» проставлен признак "
            "«умножается наценкой за размер» — сегодня одно полотно."
        ),
        (
            "Перемножается с коэффициентом формы: круглое 2300 мм "
            "получит обе наценки."
        ),
    )
    for offset, rule in enumerate(rules):
        cell = sheet.cell(
            row=DATA_START + len(PLANNED_SURCHARGE) + 2 + offset,
            column=1,
            value=f"• {rule}",
        )
        cell.font = Font(size=10, color=SOFT_INK)
    _widths(sheet, {"A": 30, "B": 16, "C": 60})
    _tip(
        sheet,
        "A4",
        "Порог сверяется с наибольшей стороной изделия, с поворотом: "
        "2300 × 900 и 900 × 2300 получат одну ступень.",
    )


def _check_sheet(
    sheet: Worksheet,
    product: Product | None,
    settings: PricingSettings | None,
) -> None:
    """Итоги движка сайта: по ним книга и сверяется.

    Числа здесь считает `memiro.pricing` — тот же расчёт, что и на
    витрине. Соберите на «Расчёте» то же зеркало: разойдутся — дефект
    в формулах книги, и чинить надо её, а не сайт.
    """
    _title(sheet, "A1", "Сверка с движком сайта")
    _subtitle(
        sheet,
        "A2",
        "Итоги посчитаны кодом сайта на разметке товара. Введите тот "
        "же размер на листе «Расчёт» — числа обязаны совпасть. "
        "Наценки за размер в движке ещё нет, поэтому сверяйте при "
        "пустой таблице ступеней.",
        "A2:E2",
    )
    if product is None:
        note = sheet.cell(
            row=4,
            column=1,
            value="Товар не назван — сверять не с чем. Запустите "
            "команду с --product <слаг>.",
        )
        note.font = Font(size=11, italic=True, color=BRASS)
        _widths(sheet, {"A": 70})
        return
    _header(
        sheet,
        4,
        (
            "Размер",
            "Площадь, м²",
            "Периметр, пог. м",
            "Итог движка, ₽",
            "Платных статей",
        ),
    )
    limits = tariffs.limits_from_settings()
    for row, size in enumerate(CHECK_SIZES, DATA_START):
        _check_row(sheet, row, product, size, limits)
    _widths(sheet, {"A": 20, "B": 14, "C": 18, "D": 16, "E": 16})
    caption = sheet.cell(
        row=DATA_START + len(CHECK_SIZES) + 1,
        column=1,
        value=f"Товар: {product.name}"
        + ("" if settings else "  ·  параметры расчёта не заведены"),
    )
    caption.font = Font(size=10, italic=True, color=SOFT_INK)


def _check_row(
    sheet: Worksheet,
    row: int,
    product: Product,
    size: tuple[int, int],
    limits: pricing.PricingLimits,
) -> None:
    """Одна строка сверки: размер и то, во что его оценил движок."""
    width, height = size
    configuration = tariffs.configuration(
        product, width_mm=width, height_mm=height
    )
    price = tariffs.price(configuration, limits=limits)
    fits = limits.fits(width_mm=width, height_mm=height)
    cells: tuple[object, ...] = (
        f"{width} × {height} мм",
        configuration.area_m2,
        configuration.perimeter_m,
        price.total if fits else "за пределом",
        len(price.lines),
    )
    for column, value in enumerate(cells, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
    sheet.cell(row=row, column=2).number_format = "0.000"
    sheet.cell(row=row, column=3).number_format = "0.000"
    total = sheet.cell(row=row, column=4)
    total.number_format = MONEY_ROUND if fits else "General"
    total.font = Font(bold=True, color=BRASS)


def write(
    category: Category,
    path: str,
    *,
    product: Product | None = None,
) -> None:
    """Собрать книгу и положить файлом."""
    build(category, product=product).save(path)
